"""
Script para extrair todas as informações da planilha BioCalc_EngS.xlsx
Extrai: abas, valores, fórmulas, tabelas auxiliares e estrutura
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import csv
import os
from pathlib import Path

# Caminho para a planilha
EXCEL_PATH = Path(__file__).parent.parent / "definições" / "BioCalc_EngS.xlsx"
OUTPUT_DIR = Path(__file__).parent.parent / "extracted_data"

def create_output_dir():
    """Cria diretório de saída se não existir"""
    OUTPUT_DIR.mkdir(exist_ok=True)
    print(f"✓ Diretório de saída criado: {OUTPUT_DIR}")

def extract_sheet_info(wb):
    """Extrai informações básicas sobre todas as abas"""
    sheets_info = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Encontrar última linha e coluna com dados
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        sheets_info.append({
            "name": sheet_name,
            "max_row": max_row,
            "max_col": max_col,
            "dimensions": f"{get_column_letter(max_col)}{max_row}"
        })
    
    # Salvar em JSON
    with open(OUTPUT_DIR / "sheets_info.json", "w", encoding="utf-8") as f:
        json.dump(sheets_info, f, indent=2, ensure_ascii=False)
    
    print(f"✓ Informações de {len(sheets_info)} abas extraídas")
    return sheets_info

def extract_sheet_data(sheet, sheet_name):
    """Extrai dados completos de uma aba (valores e fórmulas)"""
    data = []
    formulas = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
        for col_idx, cell in enumerate(row, start=1):
            col_letter = get_column_letter(col_idx)
            cell_ref = f"{col_letter}{row_idx}"
            
            # Valor da célula
            value = cell.value
            
            # Fórmula (se existir)
            formula = None
            if cell.data_type == 'f':  # É uma fórmula
                formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
            
            # Formatação e estilo
            fill_color = None
            if cell.fill and cell.fill.start_color:
                # Converter RGB para string para serialização JSON
                rgb = cell.fill.start_color.rgb
                fill_color = str(rgb) if rgb else None
            
            # Só adicionar se tiver conteúdo
            if value is not None or formula is not None:
                cell_data = {
                    "cell": cell_ref,
                    "row": row_idx,
                    "col": col_idx,
                    "value": str(value) if value is not None else None,
                    "type": cell.data_type,
                    "fill_color": fill_color
                }
                
                if formula:
                    cell_data["formula"] = formula
                    formulas.append({
                        "cell": cell_ref,
                        "formula": formula,
                        "result": str(value) if value is not None else None
                    })
                
                data.append(cell_data)
    
    # Salvar dados da aba
    with open(OUTPUT_DIR / f"sheet_{sheet_name}_data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    # Salvar fórmulas separadamente
    if formulas:
        with open(OUTPUT_DIR / f"sheet_{sheet_name}_formulas.json", "w", encoding="utf-8") as f:
            json.dump(formulas, f, indent=2, ensure_ascii=False)
        
        # Também em CSV para fácil visualização
        with open(OUTPUT_DIR / f"sheet_{sheet_name}_formulas.csv", "w", encoding="utf-8", newline='') as f:
            writer = csv.DictWriter(f, fieldnames=["cell", "formula", "result"])
            writer.writeheader()
            writer.writerows(formulas)
    
    print(f"  ✓ Aba '{sheet_name}': {len(data)} células, {len(formulas)} fórmulas")
    return data, formulas

def identify_tables(data, sheet_name):
    """Identifica possíveis tabelas baseado em padrões de dados"""
    # Agrupar células por linha
    rows = {}
    for cell in data:
        row_num = cell["row"]
        if row_num not in rows:
            rows[row_num] = []
        rows[row_num].append(cell)
    
    # Procurar por linhas consecutivas com dados (possíveis tabelas)
    tables = []
    current_table = []
    
    for row_num in sorted(rows.keys()):
        if len(rows[row_num]) >= 2:  # Pelo menos 2 células na linha
            current_table.append(row_num)
        else:
            if len(current_table) >= 3:  # Tabela com pelo menos 3 linhas
                tables.append({
                    "start_row": current_table[0],
                    "end_row": current_table[-1],
                    "num_rows": len(current_table)
                })
            current_table = []
    
    # Verificar última tabela
    if len(current_table) >= 3:
        tables.append({
            "start_row": current_table[0],
            "end_row": current_table[-1],
            "num_rows": len(current_table)
        })
    
    if tables:
        with open(OUTPUT_DIR / f"sheet_{sheet_name}_tables.json", "w", encoding="utf-8") as f:
            json.dump(tables, f, indent=2, ensure_ascii=False)
        print(f"  ✓ {len(tables)} possíveis tabelas identificadas")
    
    return tables

def extract_colored_cells(data, sheet_name):
    """Extrai células com cores específicas (verde = input, azul = calculado)"""
    colored_cells = {
        "green_cells": [],  # Células de entrada
        "blue_cells": [],   # Células calculadas
        "other_colors": []
    }
    
    for cell in data:
        if cell.get("fill_color"):
            color = cell["fill_color"]
            cell_info = {
                "cell": cell["cell"],
                "value": cell["value"],
                "color": color
            }
            
            # Verde claro (entrada de dados)
            if color and ("C6EFCE" in color or "E2EFDA" in color or "92D050" in color):
                colored_cells["green_cells"].append(cell_info)
            # Azul claro (calculado/automático)
            elif color and ("9BC2E6" in color or "BDD7EE" in color or "DDEBF7" in color):
                colored_cells["blue_cells"].append(cell_info)
            else:
                colored_cells["other_colors"].append(cell_info)
    
    if any(colored_cells.values()):
        with open(OUTPUT_DIR / f"sheet_{sheet_name}_colored_cells.json", "w", encoding="utf-8") as f:
            json.dump(colored_cells, f, indent=2, ensure_ascii=False)
        
        print(f"  ✓ Células coloridas: {len(colored_cells['green_cells'])} verdes (input), "
              f"{len(colored_cells['blue_cells'])} azuis (calc)")
    
    return colored_cells

def create_summary_report(sheets_info, all_formulas):
    """Cria relatório resumido em texto"""
    report_lines = []
    report_lines.append("=" * 80)
    report_lines.append("RELATÓRIO DE EXTRAÇÃO - BioCalc_EngS.xlsx")
    report_lines.append("=" * 80)
    report_lines.append("")
    
    report_lines.append(f"Total de abas: {len(sheets_info)}")
    report_lines.append("")
    
    for sheet in sheets_info:
        report_lines.append(f"Aba: {sheet['name']}")
        report_lines.append(f"  Dimensões: {sheet['dimensions']} ({sheet['max_row']} linhas x {sheet['max_col']} colunas)")
        
        if sheet['name'] in all_formulas:
            report_lines.append(f"  Fórmulas encontradas: {len(all_formulas[sheet['name']])}")
        report_lines.append("")
    
    report_lines.append("=" * 80)
    report_lines.append("ARQUIVOS GERADOS:")
    report_lines.append("=" * 80)
    report_lines.append("")
    report_lines.append("1. sheets_info.json - Informações gerais de todas as abas")
    report_lines.append("2. sheet_*_data.json - Dados completos de cada aba")
    report_lines.append("3. sheet_*_formulas.json/csv - Fórmulas de cada aba")
    report_lines.append("4. sheet_*_tables.json - Tabelas identificadas")
    report_lines.append("5. sheet_*_colored_cells.json - Células coloridas (inputs/cálculos)")
    report_lines.append("")
    report_lines.append("=" * 80)
    
    report_text = "\n".join(report_lines)
    
    with open(OUTPUT_DIR / "LEIA_ME.txt", "w", encoding="utf-8") as f:
        f.write(report_text)
    
    print("\n" + report_text)

def main():
    print("=" * 80)
    print("EXTRATOR DE DADOS - BioCalc_EngS.xlsx")
    print("=" * 80)
    print()
    
    # Verificar se arquivo existe
    if not EXCEL_PATH.exists():
        print(f"❌ Erro: Arquivo não encontrado em {EXCEL_PATH}")
        return
    
    print(f"📊 Abrindo planilha: {EXCEL_PATH.name}")
    
    # Criar diretório de saída
    create_output_dir()
    
    # Abrir planilha
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    print(f"✓ Planilha carregada com sucesso")
    print()
    
    # Extrair informações das abas
    print("📋 Extraindo informações das abas...")
    sheets_info = extract_sheet_info(wb)
    print()
    
    # Processar cada aba
    print("📊 Processando cada aba...")
    all_formulas = {}
    
    for sheet_name in wb.sheetnames:
        print(f"\n🔍 Processando aba: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Extrair dados e fórmulas
        data, formulas = extract_sheet_data(sheet, sheet_name)
        all_formulas[sheet_name] = formulas
        
        # Identificar tabelas
        identify_tables(data, sheet_name)
        
        # Extrair células coloridas
        extract_colored_cells(data, sheet_name)
    
    # Criar relatório resumido
    print("\n📝 Criando relatório resumido...")
    create_summary_report(sheets_info, all_formulas)
    
    print("\n" + "=" * 80)
    print("✅ EXTRAÇÃO CONCLUÍDA COM SUCESSO!")
    print("=" * 80)
    print(f"\n📁 Todos os arquivos foram salvos em: {OUTPUT_DIR}")
    print("\n💡 Próximos passos:")
    print("   1. Revise o arquivo LEIA_ME.txt")
    print("   2. Analise os arquivos JSON gerados")
    print("   3. Compartilhe a pasta 'extracted_data' para análise")

if __name__ == "__main__":
    main()
